from __future__ import annotations

from pathlib import Path
from typing import Dict, List, Any

from openpyxl import load_workbook

from ..config import get_project_root
from ..domain.invoice import Invoice


class ExcelService:
    """
    Invoice オブジェクトの一覧をテンプレート Excel に反映し、
    出力ファイルパスを返すサービス。

    - 法人名（UIから入力）を B1 に書き込む
    - 月ごとの kWh（○月値）を B21〜M21 に書き込む
    """

    def __init__(self, cfg: Dict[str, Any]) -> None:
        self.cfg = cfg
        self.project_root: Path = get_project_root()

    def write_invoices(self, invoices: List[Invoice], corp_name: str = "") -> str:
        """
        template_output.xlsx をベースに:
          - corp_name があれば B1 に書き込む
          - invoices の各 Invoice から 1〜12月の「○月値」を集めて
            B21〜M21 に書き込む。

        ※ 同じ月が複数あった場合、後のものが上書きされる仕様。
        """
        template_path = self.project_root / "template_output.xlsx"
        if not template_path.exists():
            return ""

        wb = load_workbook(template_path)
        sheet_name = self.cfg.get("excel_cell_map", {}).get("sheet", wb.sheetnames[0])
        ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active

        # ★ 法人名が入力されていれば B1 に書き込む
        corp_name = (corp_name or "").strip()
        if corp_name:
            # 必要なら config.json の "法人名" マッピングを使ってもよい
            # cell = self.cfg.get("excel_cell_map", {}).get("法人名", "B1")
            # ws[cell] = corp_name
            ws["B1"] = corp_name

        # --- 月ごとの値をB21〜M21に代入 ---
        month_cells = {
            1: "B21",
            2: "C21",
            3: "D21",
            4: "E21",
            5: "F21",
            6: "G21",
            7: "H21",
            8: "I21",
            9: "J21",
            10: "K21",
            11: "L21",
            12: "M21",
        }

        for invoice in invoices:
            fields = invoice.fields
            for m in range(1, 13):
                key = f"{m}月値"
                if key in fields and month_cells.get(m):
                    ws[month_cells[m]] = fields[key]

        # ★ template_output.xlsx をそのまま上書き
        out_path = self.project_root / "template_output.xlsx"
        wb.save(out_path)
        return str(out_path)
